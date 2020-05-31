# -*- coding: utf-8 -*-
import openpyxl
import re
import unicodedata
import string
import time
import itertools
import multiprocessing

def pega_tweets ( aba_planilha_tweet ):
    matrix_tweets = []

    # Retorna uma lista de palavras normalizada.
    def normaliza_tweet ( tweet_original ):
        # Elimina nomes de @usuário, #hashtags e "RT".
        limpo = ' '.join (
            re.sub ( "(@[A-Za-z0-9]+)|(#[A-Za-z0-9]+)|(\w+:\/\/\S+)|(RT)", " ", tweet_original ).split () )

        # Substitui letras com acento, elimina emoticons e converte tudo em minúsculas.
        limpo = ''.join (
            (frase for frase in unicodedata.normalize ( 'NFD', limpo ) if
             unicodedata.category ( frase ) != 'Mn') ).encode (
            'ascii', 'ignore' ).decode ( 'ascii' ).lower ()

        # Elimina pontuação
        dep = str.maketrans ( '', '', string.punctuation )
        limpo = limpo.translate ( dep ).split ()

        return limpo

    # Recupera todos os tweets de uma planilha e os salva em uma matriz.

    num_linhas = aba_planilha_tweet.max_row + 1

    for linha_t in range ( 2, num_linhas ):
        # Pega o tweet da planilha depura o texto
        texto = str ( aba_planilha_tweet.cell ( row = linha_t, column = 1 ).value )
        matrix_tweets.append(normaliza_tweet(texto))

    return matrix_tweets

def pega_keyword_frame ( nome_planilha_frames ):
    base_frames = openpyxl.load_workbook ( nome_planilha_frames )
    frames = base_frames['kw_frames']
    matriz_frames = []

    num_linhas = frames.max_row + 1
    num_colunas = frames.max_column + 1

    for linha in range ( 1, num_linhas ):
        lista_kw = []
        for coluna in range ( 2, num_colunas ):
            keyword = str ( frames.cell ( row = linha, column = coluna ).value )
            if keyword != 'None':
                lista_kw.append ( keyword )
            else:
                break
        matriz_frames.append ( lista_kw )

    return matriz_frames

def busca_frame ( tweet_normalizado, tipo_frame ):

    achados = []

    for elemento in tweet_normalizado:
        for keyword in tipo_frame:
            if elemento.startswith ( keyword ):
                achados.append ( elemento )
    return achados

nome_base_tweets_xls = "identificador_de_frames.xlsx"
base_tweets = openpyxl.load_workbook ( nome_base_tweets_xls )
aba_tweets = base_tweets [ "tweets" ]

tipo_frame = pega_keyword_frame ( "keyword_frames_horiz.xlsx" )

contador = time.time ()

tweets = pega_tweets ( aba_tweets )

timer_pega_tweets = time.time () - contador

num_tweets = len ( tweets )
num_frames = 5

for t in range ( 0, num_tweets ):
    print ( "\nTWEET [{}]: {}".format ( t + 1, tweets[t] ) )

    for f in range ( 0, num_frames ):
        resultado = busca_frame ( tweets[t], tipo_frame[f] )

        if resultado:
            aba_tweets.cell ( t+2, f+2, value = ", ".join ( resultado ) )
            print ( "\tDO TIPO {} ENCONTRADOS: {}".format ( f + 1, resultado ) )

base_tweets.save ( filename = nome_base_tweets_xls )

tempo = time.time () - contador
print ( "\n-TEMPO TOTAL DE PRECESSAMENTO: {} segundos".format ( int ( tempo ) ) )
print ( "\t\t+ SIMULAÇÃO PARA 350 MIL TWEETS:" )
print ( "\t\t\t- PEGAR TWEETS: {:.1f} minutos ---".format ( timer_pega_tweets * 100 / 60 ) )
print ( "\t\t\t- TEMPO TOTAL: {:.1f} horas".format ( (time.time () - contador) * 100 / 60 / 60 ) )
print ( "\t\t+ SIMULAÇÃO PARA 3,5 MILHÕES TWEETS:" )
print ( "\t\t\t- PEGAR TWEETS: {:.1f} horas ---".format ( timer_pega_tweets * 1000 / 60 ) )
print ( "\t\t\t- TEMPO TOTAL: {:.1f} dias".format ( (time.time () - contador) * 1000 / 60 / 60 ) )
