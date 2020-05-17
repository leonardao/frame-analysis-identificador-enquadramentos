import openpyxl
import re
import unicodedata
import string

def limpa_tweet ( tweet_original ):

    # Elimina nomes de @usuário, #hashtags e "RT".
    limpo = ' '.join ( re.sub ( "(@[A-Za-z0-9]+)|(#[A-Za-z0-9]+)|(\w+:\/\/\S+)|(RT)", " ", tweet_original ).split () )

    # Elimina letras com acento e converte tudo em minúsculas.
    limpo = ''.join ( ( frase for frase in unicodedata.normalize ( 'NFD', limpo ) if unicodedata.category ( frase ) != 'Mn') ).lower ()

    # Elinima pontuação e substitui por espaço.
    for i in limpo:
        if i in string.punctuation:
            limpo = limpo.replace ( i, " " )

    # Elimina espaço repetido entre palavras
    limpo = limpo.replace ( "   ", " " )
    limpo = " " + limpo.replace ( "  ", " " )

    return limpo

# Lê as planilha XLSX.
base_tweets = openpyxl.load_workbook ( "base_tweets.xlsx" )
keyword_frames = openpyxl.load_workbook ( "keyword_frames.xlsx" )

# Lê as duas abas "tweets" e "kw_frames" dos arquivos XLSX.
tweets = base_tweets ['tweets']
kw_frames = keyword_frames ['kw_frames']

# ciclo que passa as linha da aba "tweets".
for linha_t in range ( 2, tweets.max_row + 1 ):

    # Recupera da planilha o tweet e limpa o texto
    texto = str ( tweets.cell ( row = linha_t, column = 1 ).value )
    print ( "TWEET [{}]: {}".format ( linha_t - 1, texto ) )
    texto = limpa_tweet ( texto )

    # ciclo que passa as colunas da aba "frame".
    for coluna_fr in range ( 1, kw_frames.max_column + 1 ):

        # ciclo que passa as linha de cada coluna da aba "frame".
        for linha_fr in range ( 2, kw_frames.max_row + 1 ):

            # Recupera da planilha o a palavra-chave e limpa o texto.
            keyword = str (kw_frames.cell ( row = linha_fr, column = coluna_fr ) .value )
            keyword = limpa_tweet ( keyword )

            # Procura o a palavra-chave no texto do tweet. Se o valor retornado por rfind()
            # é >=0, significa que a palavra foi encontrata. Caso contrário, a função
            # retorna o valor -1.
            if texto.rfind(keyword) >= 0:

                # O frame foi identificado, então carrega o valor "1" na planilha,
                # que equivale a "ESTE FRAME CONSTA NO TWEET".
                tweets.cell ( linha_t, coluna_fr + 1, value = keyword )
                print ( "PALAVRA-CHAVE:{}\n" .format( keyword ) )

                # Encerra o ciclo para o conjunto de palavras-chave deste frame
                # pois o ciclo recomeçará para outro frame sobre o mesmo tweet.
                coluna_fr = coluna_fr + 1
                break

            # Controle para encerrar o ciclo anticipadamente caso todos os tweets já estejam classificados.
            if linha_t == tweets.max_row + 1:
                break
            else:
                # Grava na planilha o valor "0" que equivale a "ESTE FRAME NÃO CONSTA NO TWEET".
                tweets.cell ( linha_t, coluna_fr + 1, value = 0 )

# Salva e fecha arquivos
base_tweets.save (filename = "base_tweets.xlsx")
keyword_frames.close ()