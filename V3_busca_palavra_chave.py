# -*- coding: utf-8 -*-
import openpyxl
import re
import unicodedata
import string
import time
import itertools
import multiprocessing

def carrega_keywords_frame ( nome_planilha_frames ):

    base_frames = openpyxl.load_workbook ( nome_planilha_frames )
    frames = base_frames['kw_frames']
    matriz_frames = []

    for linha in range ( 1, frames.max_row + 1 ):
        lin = []
        for coluna in range ( 2, frames.max_column + 1 ):
            lin.append ( str ( frames.cell ( row = linha, column = coluna ).value ) )
        matriz_frames.append ( lin )

    return matriz_frames

def normaliza_tweet ( tweet_original ):

    # Elimina nomes de @usuário, #hashtags e "RT".
    limpo = ' '.join (re.sub ( "(@[A-Za-z0-9]+)|(#[A-Za-z0-9]+)|(\w+:\/\/\S+)|(RT)", " ", tweet_original ).split () )

    # Substitui letras com acento, elimina emoticons e converte tudo em minúsculas.
    limpo = ''.join (
        (frase for frase in unicodedata.normalize ( 'NFD', limpo ) if unicodedata.category ( frase ) != 'Mn') ).encode (
        'ascii', 'ignore' ).decode ( 'ascii' ).lower ()

    # Elimina pontuação
    dep = str.maketrans ( '', '', string.punctuation )
    limpo = limpo.translate(dep).split()

    return limpo

def pega_tweets ( nome_planilha_tweets ):
    base_tweets = openpyxl.load_workbook ( nome_planilha_tweets )
    tweets = base_tweets['tweets']
    matrix_tweets = []

    for linha_t in range ( 2, tweets.max_row + 1 ):
        # Pega o tweet da planilha depura o texto e o transforma em uma lista de palavras.
        texto = str ( tweets.cell ( row = linha_t, column = 1 ).value )
        matrix_tweets.append(normaliza_tweet(texto))

    return matrix_tweets

def busca_frame ( *tweet_normalizado ):

    tipo_frame = carrega_keywords_frame ( "keyword_frames_horiz.xlsx" )
    classificados = []

    for frame in range ( 0, len ( tipo_frame ) ):
        resultado = []
        comb = list ( itertools.product ( tipo_frame [frame], tweet_normalizado ) )

        for item in comb:
            if re.match ( item [0], item [1] ):
                resultado.append ( item [1] )
        if resultado:
            classificados.append( ', '.join ( resultado ) )
        comb = []
    return classificados

def carrega_tweets ( resultado, nome_planilha_tweets ):
    base_tweets = openpyxl.load_workbook ( nome_planilha_tweets )
    tweets = base_tweets['tweets']

    for linha_t in range ( 2, tweets.max_row + 1 ):
        tweets.cell ( linha_t, 3, value = ", ".join ( resultado ) )

tipo_frame = carrega_keywords_frame ( "keyword_frames_horiz.xlsx" )

contador = time.time ()

tweets = pega_tweets ( "TESTE-identificador_de_frames.xlsx" )

resultado = []

for tweet in itertools.starmap ( busca_frame, tweets ):
    resultado.append ( list ( tweet ) )
    print ( "\nTWEET [{}]: {}".format ( len ( resultado ), tweets [ len ( resultado ) - 1 ] ) )
    print ( "\nIDENTIFICADORES ENCONTRADOS: {}".format ( resultado [ len ( resultado ) - 1 ] ))

print ( "\n--- RESULTADO: {:.2f} segundos ---".format ( time.time () - contador ) )
print ( "--- SIMULAÇÃO: {:.2f} minutos ---".format ( (time.time () - contador) * 1.67 ) )