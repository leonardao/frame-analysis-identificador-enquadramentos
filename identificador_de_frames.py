# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import re
import unicodedata
import string
import csv

def pega_tweets ( aba_planilha_tweet ):
    matrix_tweets = []
    # Retorna uma lista de palavras normalizada.
    def normaliza_tweet ( tweet_original ):
        # Elimina nomes de @usuário, #hashtags e "RT".
        limpo = ' '.join (
            re.sub ( "(@[A-Za-z0-9]+)|(#[A-Za-z0-9]+)|(\w+:\/\/\S+)|(RT)", " ", tweet_original ).split () )

        # Substitui letras com acento, elimina emoticons e converte tudo em minúsculas.
        limpo = ''.join (
            (frase for frase in unicodedata.normalize ( 'NFD', limpo ) if
             unicodedata.category ( frase ) != 'Mn') ).encode (
            'ascii', 'ignore' ).decode ( 'ascii' ).lower ()

        # Elimina pontuação
        dep = str.maketrans ( '', '', string.punctuation )
        limpo = limpo.translate ( dep )
        limpo = ' ' + limpo + ' '
        return limpo

    # Recupera todos os tweets de uma planilha e os salva em uma matriz.
    num_linhas = aba_planilha_tweet.max_row + 1

    for linha_t in range ( 2, num_linhas ):
        # Pega o tweet da planilha depura o texto
        texto = str ( aba_planilha_tweet.cell ( row = linha_t, column = 1 ).value )
        matrix_tweets.append(normaliza_tweet(texto))

    return matrix_tweets

def pega_keyword_frame ( nome_base_frames ):
    with open ( nome_base_frames ) as csv_file:
        base_frames = csv.reader ( csv_file, delimiter = ',' )
        matriz_frames = []
        for linha in base_frames:
            list ( linha.pop ( 0 ) )
            matriz_frames.append ( linha )
    return matriz_frames

def busca_frame ( tweet_normalizado, tipo_frame ):
    achados = []
    for keyword in tipo_frame:
        if keyword in tweet_normalizado:
            achados.append ( keyword )
    return achados

nome_base_tweets_xlsx = "2b-base_tweets_classificada.xlsx"
base_tweets = load_workbook ( nome_base_tweets_xlsx )
aba_tweets = base_tweets [ "tweets" ]

tipo_frame = pega_keyword_frame ( "1-base_keywords_frames.csv" )

tweets = pega_tweets ( aba_tweets )

num_tweets = len ( tweets )
num_frames = 6
frames_genericos_especificos = [ 'ATRIBUIÇÃO DE RESPONSABILIDADE', 'CONFLITO', 'MORALIDADE', 'CONSEQUÊNCIAS DA PANDEMIA', 'MEDIDAS DE CONTENÇÃO', 'MÉTODOS DE TRATAMENTO' ]

for t in range ( 0, num_tweets ):
    #print ( "\nTWEET [{}]:{}".format ( t + 1, tweets[t] ) )

    for f in range ( 0, num_frames ):
        resultado = busca_frame ( tweets[t], tipo_frame[f] )

        if resultado:
            aba_tweets.cell ( t + 2, f + 2, value = ", ".join ( resultado ) )
			#aba_tweets.cell ( t + 2, f + 2, value = "1" )
            #print ( "\t{}: {}".format ( frames_genericos_especificos[f], resultado ) )
        else:
            aba_tweets.cell ( t + 2, f + 2, value = "0" )
			
base_tweets.save ( filename = nome_base_tweets_xlsx )